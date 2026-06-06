"""验证 V3 财务模型数据"""
import openpyxl

wb = openpyxl.load_workbook('D:/shFintech/QuantInsight_Pro_Financial_Model_V3.xlsx', data_only=False)
print(f'Sheets ({len(wb.sheetnames)}): {wb.sheetnames}')
print()

# 关键校验: 营收预测
ws = wb['5年营收预测_基准']
print('=== 营收预测_基准 ===')
for r in range(1, ws.max_row+1):
    row = [ws.cell(row=r, column=c).value for c in range(1, 9)]
    if any(v is not None for v in row):
        print(f'  R{r}: {row}')

print()
# 关键校验: 成本与利润
ws = wb['成本与利润_基准']
print('=== 成本与利润_基准 ===')
for r in range(1, 30):
    row = [ws.cell(row=r, column=c).value for c in range(1, 9)]
    if any(v is not None for v in row):
        print(f'  R{r}: {row}')
