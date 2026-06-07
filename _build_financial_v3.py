"""
T15 财务模型 V3 - 资管科技子公司基准
9 大客户分群, 4 轮融资, 5 场景敏感性, 单位经济, 估值路径
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============== 样式 ==============
TITLE_FONT = Font(name='SimHei', size=14, bold=True, color='FFFFFF')
HEADER_FONT = Font(name='SimHei', size=10, bold=True, color='FFFFFF')
NORMAL_FONT = Font(name='SimHei', size=10, bold=False)
BOLD_FONT = Font(name='SimHei', size=10, bold=True)
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FILL = PatternFill('solid', fgColor='2E86AB')
LIGHT_FILL = PatternFill('solid', fgColor='D6E9F8')
HIGHLIGHT_FILL = PatternFill('solid', fgColor='FFE699')
THIN = Side(border_style='thin', color='999999')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

def write_title(ws, row, col, text, span=6):
    ws.cell(row=row, column=col, value=text)
    cell = ws.cell(row=row, column=col)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = CENTER
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
        for c in range(col+1, col+span):
            ws.cell(row=row, column=c).fill = TITLE_FILL

def write_header(ws, row, col, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col+i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

def write_data(ws, row, col, data, fmt=None, bold=False, highlight=False):
    for i, v in enumerate(data):
        c = ws.cell(row=row, column=col+i, value=v)
        c.font = BOLD_FONT if bold else NORMAL_FONT
        c.border = BORDER
        c.alignment = RIGHT if isinstance(v, (int, float)) else LEFT
        if fmt:
            c.number_format = fmt
        if highlight:
            c.fill = HIGHLIGHT_FILL

# ============== Sheet 1: 封面与摘要 ==============
ws = wb.create_sheet('封面与摘要')
ws.column_dimensions['A'].width = 28
for col in 'BCDEFG':
    ws.column_dimensions[col].width = 16

ws['A1'] = 'QuantInsight Pro 财务模型 V3.0 - 资管科技子公司基准'
ws['A1'].font = Font(name='SimHei', size=18, bold=True, color='1F4E78')
ws.merge_cells('A1:G1')
ws['A1'].alignment = CENTER

rows = [
    ('编制单位', '慧点资本 (InsightQuant) 量化研究部'),
    ('项目主体', '杭州永字资产管理有限公司 资管科技子公司 (拟设立)'),
    ('推荐单位', '杭州永字资产管理有限公司'),
    ('项目编号', '2026FINTECH-FINT-0093'),
    ('预测期间', '2026-2030 年 (5 年)'),
    ('数据基准', '业内资管科技 SaaS 同业 (Wind/同花顺 iFinD/恒生聚源/通联数据)'),
    ('货币单位', '人民币 (万元)'),
    ('模型版本', 'V3.0 (2026-06-06 重写)'),
    ('', ''),
    ('== V3 相对 V2 改进 ==', ''),
    ('1. 客户分群从 4 类扩到 9 类', '更精细定价 (10万-500万/年/家)'),
    ('2. 新增融资轮次路径 (天使/Pre-A/A/B/Pre-IPO)', '5 年累计 4.8 亿 (V2 仅 0.18 亿)'),
    ('3. 新增单位经济分析 (CAC/LTV/回收期)', 'CAC 8-15 万, LTV 80-150 万, 回收期 6-12 月'),
    ('4. 新增客户成功指标 (续约率/NRR/流失率)', '续约率 92%, NRR 110-120%, 流失率 8%'),
    ('5. 新增 5 场景敏感性 (牛/基/熊/极端牛/极端熊)', '基线 IRR 28%, 熊市 5%, 牛市 45%'),
    ('6. 新增估值路径 (DCF/可比/风险溢价)', '第 5 年末估值 30-50 亿 (10-12x P/S)'),
    ('7. 新增里程碑节点 (12/24/36/48/60 月)', '对应营收/客户/团队/估值'),
    ('8. 团队从 80 人扩到 100 人, 销售比从 18% 升到 25%', '匹配 B 轮后销售扩张需求'),
]
for i, (k, v) in enumerate(rows, start=3):
    ws.cell(row=i, column=1, value=k).font = BOLD_FONT if k else NORMAL_FONT
    ws.cell(row=i, column=2, value=v).font = NORMAL_FONT
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
    ws.cell(row=i, column=2).alignment = LEFT

# ============== Sheet 2: 9 类客户分群与定价 ==============
ws = wb.create_sheet('客户分群与定价')
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 22
for col in 'CDEFGHIJ':
    ws.column_dimensions[col].width = 14

write_title(ws, 1, 1, '9 类客户分群与定价矩阵 (2026 年价)', span=10)
write_header(ws, 2, 1, ['客户类型', '代表客户', '年费 (万/家)', '决策周期', '决策人', 'LTV (年)', 'CAC (万)', '目标客户数', '市场容量', '覆盖策略'])

client_data = [
    ('中小私募', '敦和/明汯/乐瑞/幻方', '50', '3-6 月', '投资总监/合伙人', '4.5', '6', '500', '3000', '直销 + 推荐单位'),
    ('银行资管', '招行/兴业/浦发理财子', '200', '12-18 月', '资管部总经理', '6', '25', '20', '30', '高层 BD + 试点'),
    ('券商资管', '国君/海通/中信证券资管', '150', '9-12 月', 'CIO/首席', '5', '20', '40', '60', '战略联盟 + 共建'),
    ('信托公司', '中信信托/平安信托', '100', '6-12 月', '证券投资部总', '4', '15', '30', '50', '行业会议 + 案例'),
    ('公募基金', '易方达/广发/华夏基金', '500', '18-24 月', '总经理 + IT', '8', '60', '15', '30', '招投标 + 高层'),
    ('保险资管', '国寿/平安/泰康资管', '300', '12-18 月', '首席投资官', '6', '35', '20', '30', '监管沙盒 + 试点'),
    ('高校研究所', '清华/北大/上财', '20', '3-6 月', '金融学院院长', '3', '3', '50', '200', '学术合作 + 免费培训'),
    ('上市公司IR', '宁德/比亚迪/招行', '30', '6-9 月', '董秘/IR 总监', '3', '5', '100', '500', '路演 + 案例'),
    ('战略合作', '推荐单位/同集团', '100', '1-3 月', '集团领导', '5', '8', '20', '20', '内部协同'),
]
for i, row in enumerate(client_data, start=3):
    write_data(ws, i, 1, row)
    # Highlight key columns
    for col in [3, 6, 7, 8]:
        ws.cell(row=i, column=col).fill = HIGHLIGHT_FILL

# 汇总行
i = 12
ws.cell(row=i, column=1, value='合计 / 加权').font = BOLD_FONT
ws.cell(row=i, column=3, value='加权 ARPU = 79 万').font = BOLD_FONT
ws.cell(row=i, column=6, value='加权 LTV = 5.1 年').font = BOLD_FONT
ws.cell(row=i, column=7, value='加权 CAC = 11 万').font = BOLD_FONT
ws.cell(row=i, column=8, value='目标客户 795 家').font = BOLD_FONT
for c in range(1, 11):
    ws.cell(row=i, column=c).fill = LIGHT_FILL

# 定价逻辑
i = 14
ws.cell(row=i, column=1, value='== 定价逻辑 ==').font = BOLD_FONT
i += 1
ws.cell(row=i, column=1, value='基础价 = 数据 + SaaS + AI 投研模块')
i += 1
ws.cell(row=i, column=1, value='+ 模块加价: NLP/因子/回测/风控/智能体 (每个 10-30 万)')
i += 1
ws.cell(row=i, column=1, value='+ 客户规模调整: AUM 100 亿以下 -30%, 1000 亿以上 +50%')
i += 1
ws.cell(row=i, column=1, value='+ 续约折扣: 续约客户 + 客户成功投入 -5% 至 +20% (按使用深度)')

# ============== Sheet 3: 5 年营收预测_基准 ==============
ws = wb.create_sheet('5年营收预测_基准')
ws.column_dimensions['A'].width = 20
for col in 'BCDEFGH':
    ws.column_dimensions[col].width = 14

write_title(ws, 1, 1, '5 年营收预测 (基准场景, 资管科技子公司基准)', span=8)
write_header(ws, 2, 1, ['项目', '2026', '2027', '2028', '2029', '2030', '5 年合计', 'CAGR'])

# 客户增长
customer_rows = [
    ('客户数 (家, 年末累计)', 30, 85, 200, 380, 620, 1315, None),
    ('新签客户 (家)', 30, 65, 130, 220, 290, 735, None),
    ('流失客户 (家, 8% 流失率)', 0, 10, 15, 40, 50, 115, None),
    ('净增客户 (家)', 30, 55, 115, 180, 240, 620, None),
    ('ARPU (万/家/年, 加权)', 50, 65, 75, 82, 90, 79, None),
]
write_data(ws, 3, 1, customer_rows[0], bold=True, highlight=True)
write_data(ws, 4, 1, customer_rows[1])
write_data(ws, 5, 1, customer_rows[2])
write_data(ws, 6, 1, customer_rows[3])
write_data(ws, 7, 1, customer_rows[4], bold=True)

# 营收 (按客户分群, 简化为加权)
rev_start = 9
ws.cell(row=rev_start, column=1, value='== 营业收入 (万元) ==').font = BOLD_FONT
ws.cell(row=rev_start, column=1).fill = LIGHT_FILL

rev_rows = [
    ('SaaS 订阅 (年付)', 1500, 5525, 15000, 31160, 55800, 108985, 147),  # 客户×ARPU×订阅比例
    ('定制开发 (项目制)', 200, 800, 2500, 6000, 12000, 21500, 178),
    ('数据 API 调用 (计量)', 150, 425, 1200, 2700, 5500, 9975, 147),
    ('投研培训 + 咨询', 50, 200, 600, 1400, 3200, 5450, 183),
    ('战略合作 (推荐单位导入)', 100, 350, 800, 1700, 3000, 5950, 134),
]
for i, row in enumerate(rev_rows, start=rev_start+1):
    write_data(ws, i, 1, row)
    if i == rev_start+1:
        ws.cell(row=i, column=3).fill = HIGHLIGHT_FILL  # 突出主收入

# 营收合计
total_rev_row = rev_start + 6
ws.cell(row=total_rev_row, column=1, value='营业总收入').font = BOLD_FONT
ws.cell(row=total_rev_row, column=1).fill = HEADER_FILL
ws.cell(row=total_rev_row, column=1).font = HEADER_FONT
totals = [2000, 7300, 20100, 42960, 79500, 151860, 147]
for c, v in enumerate(totals, start=2):
    cell = ws.cell(row=total_rev_row, column=c, value=v)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = RIGHT
    cell.border = BORDER

# 同比增长率
i = total_rev_row + 1
ws.cell(row=i, column=1, value='同比增长 (%)')
ws.cell(row=i, column=2, value='-')
ws.cell(row=i, column=3, value='265%')
ws.cell(row=i, column=4, value='175%')
ws.cell(row=i, column=5, value='114%')
ws.cell(row=i, column=6, value='85%')
ws.cell(row=i, column=8, value='147% CAGR')
for c in range(1, 9):
    ws.cell(row=i, column=c).font = NORMAL_FONT
    ws.cell(row=i, column=c).alignment = RIGHT

# 收入结构
i = total_rev_row + 3
ws.cell(row=i, column=1, value='== 收入结构 (基准年 2026) ==').font = BOLD_FONT
struct_rows = [
    ('SaaS 订阅', '75%', '主要收入来源, 稳定可预期'),
    ('定制开发', '10%', '高毛利, 提升大客户粘性'),
    ('数据 API', '7.5%', '补充, 拓展新场景'),
    ('投研培训', '2.5%', '品牌建设 + 客户教育'),
    ('战略合作', '5%', '推荐单位导入, 示范效应'),
]
for j, (k, p, d) in enumerate(struct_rows, start=i+1):
    ws.cell(row=j, column=1, value=k).font = NORMAL_FONT
    ws.cell(row=j, column=2, value=p).font = BOLD_FONT
    ws.cell(row=j, column=3, value=d).font = NORMAL_FONT
    ws.merge_cells(start_row=j, start_column=3, end_row=j, end_column=8)

# ============== Sheet 4: 成本与利润_基准 ==============
ws = wb.create_sheet('成本与利润_基准')
ws.column_dimensions['A'].width = 22
for col in 'BCDEFGH':
    ws.column_dimensions[col].width = 14

write_title(ws, 1, 1, '5 年成本与利润预测 (基准场景)', span=8)
write_header(ws, 2, 1, ['项目', '2026', '2027', '2028', '2029', '2030', '5 年合计', '占比/率'])

# 营业成本
cost_rows = [
    ('== 营业成本 ==', None, None, None, None, None, None, None),
    ('数据采购', 100, 290, 600, 1100, 1800, 3890, '4.5% 营收'),
    ('云服务 / 算力', 150, 380, 700, 1300, 2200, 4730, '5% 营收'),
    ('第三方 AI 接口 (DeepSeek/Qwen)', 80, 200, 350, 600, 900, 2130, '2.5%'),
    ('人力 (研发 60%)', 800, 1800, 3200, 5500, 8500, 19800, '技术岗 + 基础'),
    ('客户成功 (续约支持)', 60, 200, 500, 1100, 2200, 4060, '5-8%'),
    ('差旅 + 销售成本', 50, 180, 450, 950, 1900, 3530, '5% 营收'),
    ('营业成本合计', 1240, 3050, 5800, 10550, 17500, 38140, '27% 营收'),
]
for i, row in enumerate(cost_rows, start=3):
    if row[0].startswith('=='):
        ws.cell(row=i, column=1, value=row[0]).font = BOLD_FONT
        ws.cell(row=i, column=1).fill = LIGHT_FILL
    elif row[0].endswith('合计'):
        write_data(ws, i, 1, row, bold=True, highlight=True)
    else:
        write_data(ws, i, 1, row)

# 毛利
gross_row = 12
ws.cell(row=gross_row, column=1, value='毛利润').font = BOLD_FONT
gross = [760, 4250, 14300, 32410, 62000, 113720, 75]
for c, v in enumerate(gross, start=2):
    cell = ws.cell(row=gross_row, column=c, value=v)
    cell.font = BOLD_FONT
    cell.alignment = RIGHT
    cell.fill = HIGHLIGHT_FILL
    cell.border = BORDER

ws.cell(row=gross_row+1, column=1, value='毛利率 (%)')
rates = [38, 58, 71, 75, 78, 75, None]
for c, v in enumerate(rates, start=2):
    if v is not None:
        ws.cell(row=gross_row+1, column=c, value=f'{v}%').font = NORMAL_FONT
    ws.cell(row=gross_row+1, column=c).alignment = RIGHT

# 期间费用
opex_start = gross_row + 3
ws.cell(row=opex_start, column=1, value='== 期间费用 ==').font = BOLD_FONT
ws.cell(row=opex_start, column=1).fill = LIGHT_FILL

opex_rows = [
    ('研发 (含 AI 训练)', 1500, 2500, 4500, 7500, 12000, 28000, '15% 营收'),
    ('销售 (含 BD)', 800, 1800, 4200, 9000, 18000, 33800, '23% 营收'),
    ('管理 (含财务/法务)', 200, 400, 800, 1500, 2800, 5700, '3.5%'),
    ('营销', 100, 350, 900, 2000, 4000, 7350, '5%'),
    ('期间费用合计', 2600, 5050, 10400, 20000, 36800, 74850, '49% 营收'),
]
for i, row in enumerate(opex_rows, start=opex_start+1):
    if row[0].endswith('合计'):
        write_data(ws, i, 1, row, bold=True, highlight=True)
    else:
        write_data(ws, i, 1, row)

# 营业利润
op_profit_row = opex_start + 6
ws.cell(row=op_profit_row, column=1, value='营业利润').font = BOLD_FONT
op_profit = [-1840, -800, 3900, 12410, 25200, 38870, 26]
for c, v in enumerate(op_profit, start=2):
    cell = ws.cell(row=op_profit_row, column=c, value=v)
    cell.font = BOLD_FONT
    cell.alignment = RIGHT
    cell.fill = HIGHLIGHT_FILL
    cell.border = BORDER

# 净利率
ws.cell(row=op_profit_row+1, column=1, value='营业利润率 (%)')
op_rates = [-92, -11, 19, 29, 32, 26, None]
for c, v in enumerate(op_rates, start=2):
    if v is not None:
        ws.cell(row=op_profit_row+1, column=c, value=f'{v}%').font = NORMAL_FONT
    ws.cell(row=op_profit_row+1, column=c).alignment = RIGHT

# 净利润
ni_row = op_profit_row + 3
ws.cell(row=ni_row, column=1, value='净利润').font = BOLD_FONT
ws.cell(row=ni_row, column=1).fill = HEADER_FILL
ws.cell(row=ni_row, column=1).font = HEADER_FONT
ni = [-1840, -800, 2925, 9310, 19800, 29395, 22]
for c, v in enumerate(ni, start=2):
    cell = ws.cell(row=ni_row, column=c, value=v)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = RIGHT
    cell.border = BORDER

ws.cell(row=ni_row+1, column=1, value='净利率 (%)')
ni_rates = [-92, -11, 15, 22, 25, 22, None]
for c, v in enumerate(ni_rates, start=2):
    if v is not None:
        ws.cell(row=ni_row+1, column=c, value=f'{v}%').font = NORMAL_FONT
    ws.cell(row=ni_row+1, column=c).alignment = RIGHT

ws.cell(row=ni_row+2, column=1, value='(适用所得税 0%/0%/15%/15%/15%)').font = NORMAL_FONT

# ============== Sheet 5: 单位经济与客户成功 ==============
ws = wb.create_sheet('单位经济与续约')
ws.column_dimensions['A'].width = 22
for col in 'BCDEFGH':
    ws.column_dimensions[col].width = 14

write_title(ws, 1, 1, '单位经济 (Unit Economics) 与客户成功指标', span=8)
write_header(ws, 2, 1, ['指标', '2026', '2027', '2028', '2029', '2030', '行业基准', '对标'])

unit_rows = [
    ('CAC (客户获取成本, 万/家)', 12, 11, 10, 9, 8, '8-15', '处于行业中位'),
    ('LTV (客户终身价值, 万/家)', 80, 130, 180, 240, 320, '80-300', 'LTV/CAC > 10'),
    ('回收期 (月)', 18, 12, 8, 6, 5, '6-18', '逐年优化'),
    ('LTV / CAC', 6.7, 11.8, 18.0, 26.7, 40.0, '> 3 健康', '>>> 3'),
    ('续约率 (%)', '95%', '93%', '92%', '92%', '92%', '85-95%', '高于行业'),
    ('NRR (净收入留存, %)', '108%', '112%', '115%', '118%', '120%', '100-120%', 'SaaS 金标准'),
    ('客户流失率 (%)', '5%', '7%', '8%', '8%', '8%', '5-15%', '行业平均'),
    ('ARPU 增长 (YoY)', '-', '30%', '15%', '9%', '10%', '10-30%', '扩张销售驱动'),
    ('单人创收 (万/人/年)', 13, 24, 40, 54, 80, '50-200', '快速逼近基准'),
    ('单人毛利 (万/人/年)', 5, 14, 28, 41, 62, '30-120', '健康水平'),
]
for i, row in enumerate(unit_rows, start=3):
    write_data(ws, i, 1, row)
    if row[0] in ('LTV / CAC', '回收期 (月)'):
        for c in [2, 3, 4, 5, 6]:
            ws.cell(row=i, column=c).fill = HIGHLIGHT_FILL

# 客户成功漏斗
i = 14
ws.cell(row=i, column=1, value='== 客户成功漏斗 (2030 年稳态) ==').font = BOLD_FONT
funnel = [
    ('线索 (Leads)', '5000'),
    ('MQL (营销合格线索)', '1500'),
    ('SQL (销售合格线索)', '600'),
    ('机会 (Opportunities)', '290'),
    ('签约 (Closed Won)', '240'),
    ('续约 (Renewed, 92%)', '221'),
    ('扩展销售 (Upsell, 25%)', '55'),
    ('流失 (Churned, 8%)', '19'),
    ('净增客户', '276'),
]
for j, (k, v) in enumerate(funnel, start=i+1):
    ws.cell(row=j, column=1, value=k).font = NORMAL_FONT
    ws.cell(row=j, column=2, value=v).font = BOLD_FONT
    ws.cell(row=j, column=2).alignment = RIGHT
    if '净增' in k or '签约' in k:
        ws.cell(row=j, column=2).fill = HIGHLIGHT_FILL

# ============== Sheet 6: 融资轮次与估值 ==============
ws = wb.create_sheet('融资与估值')
ws.column_dimensions['A'].width = 22
for col in 'BCDEFGH':
    ws.column_dimensions['col'].width = 14 if col != 'A' else 22
for col in 'BCDEFGH':
    ws.column_dimensions[col].width = 16

write_title(ws, 1, 1, '5 年融资轮次 + 估值路径 (资管科技子公司视角)', span=8)
write_header(ws, 2, 1, ['轮次', '时点 (月)', '金额 (万)', '投后估值 (万)', '股本稀释', 'P/S 倍数', '领投方', '用途'])

rounds = [
    ('天使轮', '6', '500', '5000', '10%', '25x', '慧点资本 + 推荐单位', 'MVP + 5 客户'),
    ('Pre-A', '12', '3000', '30000', '10%', '20x', '一线 VC', '产品 2.0 + 30 客户'),
    ('A 轮', '24', '10000', '80000', '12.5%', '11x', '一线 VC + 产业资本', '销售扩张 + 100 客户'),
    ('B 轮', '36', '30000', '200000', '15%', '10x', '顶级 PE + 战投', '技术壁垒 + 200 客户'),
    ('Pre-IPO', '54', '0', '500000', '0%', '8x', '二级市场', 'IPO 准备'),
    ('5 年累计融资', '', '43500', '-', '47.5%', '-', '-', '-'),
]
for i, row in enumerate(rounds, start=3):
    if row[0] == '5 年累计融资':
        write_data(ws, i, 1, row, bold=True, highlight=True)
    else:
        write_data(ws, i, 1, row)

# 估值路径
i = 10
ws.cell(row=i, column=1, value='== 估值方法 (3 种交叉验证) ==').font = BOLD_FONT

val_rows = [
    ('DCF 估值 (WACC 12%, 永续增长 3%)', '35-40 亿', '基于基准场景现金流贴现'),
    ('可比公司 P/S (Wind/同花顺/恒生)', '12-15x', '行业 2026 年中位数'),
    ('可比公司 P/E (成熟期 30x)', '45-60 亿', '假设 2030 年净利 2 亿'),
    ('风险溢价 (资管科技 β 1.3)', '+5-8%', '相对沪深 300 指数'),
    ('综合估值区间', '30-50 亿', '第 5 年末合理估值'),
]
for j, (k, v, d) in enumerate(val_rows, start=i+1):
    ws.cell(row=j, column=1, value=k).font = NORMAL_FONT
    ws.cell(row=j, column=2, value=v).font = BOLD_FONT
    ws.cell(row=j, column=2).alignment = RIGHT
    ws.cell(row=j, column=2).fill = HIGHLIGHT_FILL
    ws.cell(row=j, column=3, value=d).font = NORMAL_FONT
    ws.merge_cells(start_row=j, start_column=3, end_row=j, end_column=8)

# 团队期权池
i = 17
ws.cell(row=i, column=1, value='== 期权池预留 ==').font = BOLD_FONT
i += 1
ws.cell(row=i, column=1, value='Pre-A 后设立期权池 10% (从创始团队出)').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='B 轮后扩到 15%').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='IPO 前 20% (含 ESOP)').font = NORMAL_FONT

# ============== Sheet 7: 团队与人员规划 ==============
ws = wb.create_sheet('团队与人员规划')
ws.column_dimensions['A'].width = 22
for col in 'BCDEFGH':
    ws.column_dimensions[col].width = 14

write_title(ws, 1, 1, '5 年团队规划 (资管科技子公司)', span=8)
write_header(ws, 2, 1, ['部门', '2026', '2027', '2028', '2029', '2030', '5 年新增', '人均成本 (万)'])

team_rows = [
    ('研发 (AI + 平台)', 12, 25, 50, 80, 100, 88, 60),
    ('数据 (采编 + ETL)', 3, 8, 15, 25, 35, 32, 40),
    ('销售 + BD', 3, 10, 25, 50, 80, 77, 50),
    ('客户成功', 1, 4, 10, 20, 35, 34, 45),
    ('市场 + 品牌', 1, 3, 5, 10, 15, 14, 40),
    ('管理 (财务/法务/HR)', 2, 4, 8, 12, 15, 13, 50),
    ('小计 (全公司)', 22, 54, 113, 197, 280, 258, 53),
    ('创始团队 (4 人, 冯亦根/薛永再/黄成选/冯思涵)', 4, 4, 4, 4, 4, 0, 0),
    ('合计', 26, 58, 117, 201, 284, 258, 53),
]
for i, row in enumerate(team_rows, start=3):
    if '小计' in row[0]:
        write_data(ws, i, 1, row, bold=True, highlight=True)
    elif row[0] == '合计':
        write_data(ws, i, 1, row, bold=True)
        for c in range(1, 9):
            ws.cell(row=i, column=c).fill = HEADER_FILL
            ws.cell(row=i, column=c).font = HEADER_FONT
    else:
        write_data(ws, i, 1, row)

# 人均指标
i = 13
ws.cell(row=i, column=1, value='== 人均指标 (2030) ==').font = BOLD_FONT
i += 1
ws.cell(row=i, column=1, value='单人创收').font = NORMAL_FONT
ws.cell(row=i, column=2, value='280 万').font = BOLD_FONT
i += 1
ws.cell(row=i, column=1, value='单人毛利').font = NORMAL_FONT
ws.cell(row=i, column=2, value='220 万').font = BOLD_FONT
i += 1
ws.cell(row=i, column=1, value='人均薪酬 (含五险一金)').font = NORMAL_FONT
ws.cell(row=i, column=2, value='53 万').font = BOLD_FONT
i += 1
ws.cell(row=i, column=1, value='薪酬占毛利比').font = NORMAL_FONT
ws.cell(row=i, column=2, value='24%').font = HIGHLIGHT_FILL

# 校招 + 社招
i = 19
ws.cell(row=i, column=1, value='== 招聘策略 ==').font = BOLD_FONT
i += 1
ws.cell(row=i, column=1, value='校招 (清华/北大/上财/复旦/交大/浙大): 研发 80%, 销售 20%').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='社招 (3-5 年经验): Wind/同花顺/恒生电子/通联数据/嘉实/南方基金').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='海外 (QS Top 30): 算法/AI 工程师 5 人').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='推荐单位借调: 薛永再 6 个月 (合规 + 风控)').font = NORMAL_FONT

# ============== Sheet 8: 5 场景敏感性分析 ==============
ws = wb.create_sheet('5场景敏感性')
ws.column_dimensions['A'].width = 22
for col in 'BCDEFGH':
    ws.column_dimensions[col].width = 14

write_title(ws, 1, 1, '5 场景敏感性分析 (2030 年关键指标)', span=8)
write_header(ws, 2, 1, ['场景', '客户数', 'ARPU (万)', '营收 (亿)', '净利率', '净利润 (亿)', '估值 (亿)', 'IRR'])

scenarios = [
    ('极端熊市 (-50% 客户)', 310, 75, 3.5, '10%', 0.35, 12, '5%'),
    ('保守 (基线 -30%)', 434, 82, 5.5, '15%', 0.83, 22, '15%'),
    ('基准 (基线)', 620, 90, 7.95, '25%', 1.98, 40, '28%'),
    ('乐观 (基线 +30%)', 806, 95, 10.3, '30%', 3.09, 60, '38%'),
    ('极端牛市 (基线 +60%)', 992, 100, 12.7, '33%', 4.19, 85, '45%'),
]
for i, row in enumerate(scenarios, start=3):
    if '基准' in row[0]:
        write_data(ws, i, 1, row, bold=True, highlight=True)
    else:
        write_data(ws, i, 1, row)

# 关键假设
i = 9
ws.cell(row=i, column=1, value='== 关键假设 ==').font = BOLD_FONT
i += 1
ws.cell(row=i, column=1, value='基准: 客户数 CAGR 113%, ARPU 12%').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='极端熊市: 客户获取下降 50%, 单价 -20%, 团队 -30%').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='极端牛市: 大客户中标率 +30%, 单价 +20%, 销售人效 +30%').font = NORMAL_FONT

# 概率加权
i = 14
ws.cell(row=i, column=1, value='== 概率加权期望 ==').font = BOLD_FONT
i += 1
ws.cell(row=i, column=1, value='概率分布 (基准/乐观/保守/极端) = 50%/25%/20%/5%').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='期望 IRR = 50%*28% + 25%*38% + 20%*15% + 5%*5% = 27%').font = BOLD_FONT
i += 1
ws.cell(row=i, column=1, value='期望估值 = 50%*40 + 25%*60 + 20%*22 + 5%*12 = 39.9 亿').font = BOLD_FONT

# ============== Sheet 9: 里程碑节点 ==============
ws = wb.create_sheet('里程碑节点')
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 22
for col in 'CDEFGH':
    ws.column_dimensions[col].width = 16

write_title(ws, 1, 1, '5 年里程碑节点 (12/24/36/48/60 月)', span=8)
write_header(ws, 2, 1, ['月数', '里程碑', '客户数', '团队', '营收 (年化)', '估值', '融资轮次', '关键事件'])

milestones = [
    ('6', 'MVP + 5 客户上线', 5, 26, '1000 万', '5 亿', '天使 500 万', '立项/团队组建'),
    ('12', '产品 2.0 + 30 客户', 30, 40, '2000 万', '3 亿', 'Pre-A 3000 万', 'A 轮准备'),
    ('24', '100 客户 + 销售团队成型', 100, 80, '7300 万', '8 亿', 'A 轮 1 亿', '客户成功团队建立'),
    ('36', '200 客户 + 行业 Top 3', 200, 130, '2.0 亿', '20 亿', 'B 轮 3 亿', '海外 5 家 + 学术合作 3 家'),
    ('48', '400 客户 + 技术壁垒', 400, 200, '4.3 亿', '35 亿', 'Pre-IPO 准备', '申请专精特新 + 数据合规认证'),
    ('60', '600+ 客户 + 行业龙头', 620, 280, '8.0 亿', '40-50 亿', '拟 IPO', '启动 IPO 申报 (科创板/北交所)'),
]
for i, row in enumerate(milestones, start=3):
    write_data(ws, i, 1, row)
    if '24' in row[0] or '60' in row[0]:
        for c in range(1, 9):
            ws.cell(row=i, column=c).fill = HIGHLIGHT_FILL

# KPI 追踪
i = 11
ws.cell(row=i, column=1, value='== 关键 KPI 追踪 (董事会月报) ==').font = BOLD_FONT
kpis = [
    'MRR (月度经常性收入) + ARR (年度经常性收入)',
    'NRR (净收入留存) > 110%',
    'CAC 回收期 < 12 个月',
    '客户健康度 (CSAT) > 4.5/5',
    '产品 NPS > 50',
    '核心模型准确率 (回测 Sharpe) > 1.0',
    '员工 NPS > 40 (eNPS)',
    '现金流跑正 (基准场景 2027 Q4)',
    '合规审计 (年度)',
    '数据安全事件 0 起',
]
for j, k in enumerate(kpis, start=i+1):
    ws.cell(row=j, column=1, value=f'KPI {j-i}').font = BOLD_FONT
    ws.cell(row=j, column=2, value=k).font = NORMAL_FONT
    ws.merge_cells(start_row=j, start_column=2, end_row=j, end_column=8)

# ============== Sheet 10: 现金流与资产负债表 ==============
ws = wb.create_sheet('现金流与资负表')
ws.column_dimensions['A'].width = 22
for col in 'BCDEFG':
    ws.column_dimensions[col].width = 14

write_title(ws, 1, 1, '5 年现金流 + 资产负债表 (基准场景)', span=7)
write_header(ws, 2, 1, ['项目', '2026', '2027', '2028', '2029', '2030', '5 年合计'])

# 现金流
cf_start = 3
ws.cell(row=cf_start, column=1, value='== 现金流量表 (万) ==').font = BOLD_FONT
ws.cell(row=cf_start, column=1).fill = LIGHT_FILL

cf_rows = [
    ('经营活动现金流', -1840, -800, 2925, 9310, 19800, 29395),
    ('折旧摊销', 80, 120, 200, 300, 450, 1150),
    ('营运资本变化', -200, -500, -1000, -2000, -3500, -7200),
    ('OCF 合计', -1960, -1180, 2125, 7610, 16750, 23345),
    ('', None, None, None, None, None, None),
    ('资本支出 (设备/系统)', -100, -200, -400, -600, -1000, -2300),
    ('ICF 合计', -100, -200, -400, -600, -1000, -2300),
    ('', None, None, None, None, None, None),
    ('股权融资', 500, 3000, 10000, 30000, 0, 43500),
    ('债务融资', 0, 0, 0, 0, 0, 0),
    ('分红', 0, 0, 0, 0, 0, 0),
    ('FCF 合计', 500, 3000, 10000, 30000, 0, 43500),
    ('', None, None, None, None, None, None),
    ('现金净变动', -1560, 1620, 11725, 37010, 15750, 64545),
    ('期初现金', 0, -1560, 60, 11785, 48795, 0),
    ('期末现金', -1560, 60, 11785, 48795, 64545, 64545),
]
for i, row in enumerate(cf_rows, start=cf_start+1):
    if not row[0]:
        ws.cell(row=i, column=1, value='').font = NORMAL_FONT
    elif '合计' in row[0] or '现金' in row[0]:
        write_data(ws, i, 1, row, bold=True, highlight=True)
    else:
        write_data(ws, i, 1, row)

# 资产负债表
bs_start = cf_start + len(cf_rows) + 3
ws.cell(row=bs_start, column=1, value='== 资产负债表 (万, 年末) ==').font = BOLD_FONT
ws.cell(row=bs_start, column=1).fill = LIGHT_FILL

bs_rows = [
    ('货币资金', -1560, 60, 11785, 48795, 64545),
    ('应收账款', 200, 800, 2200, 5000, 9500),
    ('预付账款', 30, 80, 200, 400, 800),
    ('流动资产合计', -1330, 940, 14185, 54195, 74845),
    ('固定资产', 100, 200, 400, 700, 1200),
    ('无形资产 (软件 + 数据)', 200, 500, 1000, 1800, 3000),
    ('资产合计', -1030, 1640, 15585, 56695, 79045),
    ('', None, None, None, None, None),
    ('应付账款', 100, 300, 700, 1500, 3000),
    ('应付薪酬', 300, 700, 1500, 3000, 5500),
    ('合同负债 (预收款)', 150, 600, 1800, 4000, 8000),
    ('流动负债合计', 550, 1600, 4000, 8500, 16500),
    ('股东权益', -1580, 40, 11585, 48195, 62545),
    ('负债 + 权益', -1030, 1640, 15585, 56695, 79045),
]
for i, row in enumerate(bs_rows, start=bs_start+1):
    if not row[0]:
        ws.cell(row=i, column=1, value='').font = NORMAL_FONT
    elif '合计' in row[0] or '权益' in row[0]:
        write_data(ws, i, 1, row, bold=True, highlight=True)
    else:
        write_data(ws, i, 1, row)

# ============== Sheet 11: 风险与压力测试 ==============
ws = wb.create_sheet('风险与压力测试')
ws.column_dimensions['A'].width = 22
for col in 'BCDEFG':
    ws.column_dimensions[col].width = 14

write_title(ws, 1, 1, '风险情景与压力测试', span=7)
write_header(ws, 2, 1, ['风险类型', '发生概率', '对营收影响', '对估值影响', '对 IRR 影响', '应对措施', '责任人'])

risks = [
    ('AI 技术迭代滞后', '20%', '-15%', '-25%', '-8pp', '持续 R&D 投入 (15% 营收); 与 DeepSeek/Qwen 合作', '黄成选'),
    ('大客户流失', '15%', '-10%', '-15%', '-5pp', '客户成功团队扩 2x; 签 3 年长约; 定制化粘性', '冯思涵'),
    ('监管收紧 (数据/AI)', '30%', '-8%', '-12%', '-3pp', '提前 6 月预研; 参与行业协会; 法务 + 合规双保险', '黄成选'),
    ('融资环境恶化', '25%', '0 (短期)', '-30%', '-5pp', '保持 18 月现金跑道; 拓展产业资本/政府引导基金', '薛永再'),
    ('核心团队流失', '15%', '-5%', '-10%', '-4pp', '期权池 15%; 关键人 4 年 vesting + 竞业', '冯亦根'),
    ('竞品价格战 (Wind 等)', '40%', '-20%', '-20%', '-6pp', '差异化定位 (中小私募 + AI); 成本领先 30%', '冯亦根'),
    ('数据源断供', '10%', '-25%', '-30%', '-12pp', '多源备份 (Wind/akshare/聚源/通联); 自建爬虫', '黄成选'),
    ('宏观经济衰退', '25%', '-15%', '-25%', '-7pp', '客户分群分散 (9 类); 银行/保险抗周期', '冯亦根'),
]
for i, row in enumerate(risks, start=3):
    write_data(ws, i, 1, row)
    if '20%' in str(row[1]) or '40%' in str(row[1]):
        for c in range(2, 8):
            ws.cell(row=i, column=c).fill = HIGHLIGHT_FILL

# 压力测试
i = 12
ws.cell(row=i, column=1, value='== 压力测试 (2030 年关键指标) ==').font = BOLD_FONT
stress = [
    ('轻度压力 (-1 个风险)', '客户 590, 净利 1.6 亿, 估值 32 亿, IRR 23%'),
    ('中度压力 (-2 个风险)', '客户 530, 净利 1.2 亿, 估值 24 亿, IRR 17%'),
    ('重度压力 (-3 个风险)', '客户 450, 净利 0.8 亿, 估值 16 亿, IRR 10%'),
    ('极重压力 (-4+ 个风险)', '客户 350, 净利 0.3 亿, 估值 9 亿, IRR 3%'),
]
for j, (k, v) in enumerate(stress, start=i+1):
    ws.cell(row=j, column=1, value=k).font = NORMAL_FONT
    ws.cell(row=j, column=2, value=v).font = BOLD_FONT
    ws.merge_cells(start_row=j, start_column=2, end_row=j, end_column=7)

# ============== Sheet 12: 关键指标基准对标 ==============
ws = wb.create_sheet('业内基准对标')
ws.column_dimensions['A'].width = 24
for col in 'BCDEFG':
    ws.column_dimensions[col].width = 16

write_title(ws, 1, 1, '核心指标 vs 业内基准 (资管科技 SaaS 同业)', span=7)
write_header(ws, 2, 1, ['指标', 'QuantInsight Pro', 'Wind (万得)', '同花顺 iFinD', '恒生电子', '通联数据', '结论'])

bench = [
    ('目标客户', '中小私募+银行/券商/保险/信托', '机构为主', '机构 + 散户', '金融机构', '机构 + 个人', '差异化定位'),
    ('ARPU (万/家/年)', 79, '80-300', '50-150', '100-500', '20-100', '中位偏上'),
    ('客户数 (累计)', 620, '5000+', '10000+', '800+', '2000+', '聚焦高质量'),
    ('毛利率 (成熟期)', '75%', '85%+', '70%', '60%', '60%', '软件主导'),
    ('净利率 (成熟期)', '25%', '35%', '25%', '18%', '12%', '健康'),
    ('CAC (万)', 8, 30, 15, 40, 10, '极低'),
    ('LTV (万)', 320, 800, 300, 1200, 200, '中等'),
    ('续约率', '92%', '95%', '85%', '90%', '80%', '高于平均'),
    ('研发占营收 (稳态)', '15%', '12%', '8%', '20%', '25%', '保持高研发'),
    ('销售占营收 (稳态)', '23%', '20%', '18%', '15%', '30%', '匹配 B 轮'),
    ('团队 (2030)', 280, '3000+', '1500+', '8000+', '500+', '轻资产'),
    ('人均创收 (万)', 280, 200, 180, 150, 150, '人效领先'),
    ('5 年累计融资 (亿)', 4.35, '已上市', '已上市', '已上市', '已上市', '合理 B 轮'),
    ('5 年末估值 (亿)', 40, '600+', '1500+', '500+', '100+', '有上行空间'),
]
for i, row in enumerate(bench, start=3):
    write_data(ws, i, 1, row)
    if i % 2 == 0:
        for c in range(1, 8):
            ws.cell(row=i, column=c).fill = LIGHT_FILL

# 差异化优势
i = 19
ws.cell(row=i, column=1, value='== 差异化优势 (vs 业内基准) ==').font = BOLD_FONT
i += 1
ws.cell(row=i, column=1, value='1. AI 投研智能体 (业内 90% 不具备)').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='2. 中小私募聚焦 (被 Wind/同花顺 忽视的蓝海, 5000+ 家)').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='3. 多数据源融合 (5 类: 财报/公告/新闻/产业链/舆情)').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='4. 私募合规 (避免大厂合规壁垒)').font = NORMAL_FONT
i += 1
ws.cell(row=i, column=1, value='5. 学术合作背书 (清华/北大/上财, 监管沙盒)').font = NORMAL_FONT

# ============== Sheet 13: 一页纸摘要 (Pitch Sheet) ==============
ws = wb.create_sheet('一页纸摘要')
ws.column_dimensions['A'].width = 28
for col in 'BCDEFG':
    ws.column_dimensions[col].width = 14

write_title(ws, 1, 1, '一页纸财务摘要 (Pitch Sheet, 评委用)', span=7)

# 关键数字
summary = [
    ('== 关键数字 (2030 目标) ==',),
    ('客户数', '620 家'),
    ('客户分群', '9 类 (中小私募 + 银行/券商/保险/信托/公募/高校/IR/战略)'),
    ('团队', '280 人'),
    ('年营收', '7.95 亿'),
    ('净利润', '1.98 亿'),
    ('净利率', '25%'),
    ('估值 (P/S 5x)', '40 亿'),
    ('IRR (5 年)', '28%'),
    ('累计融资', '4.35 亿'),
    ('回收期', '5 年'),
    ('',),
    ('== 5 年里程碑 ==',),
    ('12 月: 30 客户 + Pre-A 3000 万', '估值 3 亿'),
    ('24 月: 100 客户 + A 轮 1 亿', '估值 8 亿'),
    ('36 月: 200 客户 + B 轮 3 亿', '估值 20 亿'),
    ('48 月: 400 客户 + Pre-IPO 准备', '估值 35 亿'),
    ('60 月: 620 客户 + 拟 IPO', '估值 40-50 亿'),
    ('',),
    ('== 单位经济金标准 ==',),
    ('LTV / CAC = 40 (>>> 3 健康)', '回收期 5 月 (远低于 18 月)'),
    ('续约率 92%, NRR 120%', '行业 Top 20%'),
    ('',),
    ('== 风险与应对 ==',),
    ('AI 滞后 (-15% 营收) → 持续 15% R&D', '✓'),
    ('大客户流失 (-10%) → 客户成功 2x + 3 年长约', '✓'),
    ('监管收紧 (-8%) → 法务 + 合规双保险 + 行业参与', '✓'),
    ('融资恶化 (-5pp IRR) → 18 月现金 + 多元融资', '✓'),
    ('',),
    ('== 评委核心关注 ==',),
    ('盈利能力: 净利率 25% 行业 Top 20%', '✓'),
    ('增长持续性: 5 年 CAGR 147%, 客户分群 9 类', '✓'),
    ('现金流: 2027 Q4 跑正, 不依赖持续融资', '✓'),
    ('估值合理性: 40 亿 5x P/S, DCF/可比/风险溢价 三角验证', '✓'),
    ('退出路径: 科创板/北交所, 60 月启动 IPO', '✓'),
]
for i, entry in enumerate(summary, start=2):
    if isinstance(entry, tuple) and len(entry) == 2:
        line, val = entry
    else:
        line = entry[0] if isinstance(entry, tuple) else entry
        val = ''
    ws.cell(row=i, column=1, value=line).font = BOLD_FONT if line.startswith('==') else NORMAL_FONT
    if '✓' in line:
        ws.cell(row=i, column=7, value=line.split(' → ')[1] if ' → ' in line else '✓').font = NORMAL_FONT
        ws.cell(row=i, column=1, value=line.split(' → ')[0]).font = NORMAL_FONT
    if '==' in line:
        ws.cell(row=i, column=1).fill = HEADER_FILL
        ws.cell(row=i, column=1).font = HEADER_FONT
    if val:
        ws.cell(row=i, column=2, value=val).font = BOLD_FONT
        ws.cell(row=i, column=2).fill = HIGHLIGHT_FILL
    if line and '==' not in line and '✓' not in line and line != '项目' and i > 2 and not val:
        if '月:' in line or '客户' in line or '估值' in line:
            ws.cell(row=i, column=2, value=line.split(' + ')[1] if ' + ' in line else '').font = NORMAL_FONT

# 保存
output_path = 'D:/shFintech/QuantInsight_Pro_Financial_Model_V3.xlsx'
wb.save(output_path)
size = os.path.getsize(output_path) / 1024
print(f'[OK] 财务模型 V3 保存: {output_path}')
print(f'[OK] 文件大小: {size:.0f} KB')
print(f'[OK] 工作表: {wb.sheetnames}')
