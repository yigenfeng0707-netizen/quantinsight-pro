"""读取 V2 财务模型的所有数据"""
import openpyxl

wb = openpyxl.load_workbook('D:/shFintech/QuantInsight_Pro_Financial_Model_V2.xlsx', data_only=False)
print(f'Sheets: {wb.sheetnames}')
print()
for name in wb.sheetnames:
    ws = wb[name]
    print(f'=== {name} ({ws.max_row} rows x {ws.max_column} cols) ===')
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        print('  ', row)
    print()
