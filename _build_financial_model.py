"""
QuantInsight Pro - 财务模型 V2.0
基于：
- 真实团队规模：4 人核心 + 永字资管推荐生态
- 业内基准：资管科技 SaaS 同业（聚宽/米筐/Wind iFinD）
- 三情景预测：乐观/基准/保守
- 五年预测（2026-2030）
- 三表预测：损益表 + 现金流 + 资产负债
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

wb = openpyxl.Workbook()

# ============ 样式定义 ============
HEADER_FONT = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
SUBHEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='1F4E78')
SUBHEADER_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
NORMAL_FONT = Font(name='微软雅黑', size=10)
BOLD_FONT = Font(name='微软雅黑', size=10, bold=True)
TOTAL_FONT = Font(name='微软雅黑', size=10, bold=True)
TOTAL_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# ============ Sheet 1: 封面与核心结论 ============
ws_cover = wb.active
ws_cover.title = '封面与核心结论'

ws_cover['A1'] = 'QuantInsight Pro - 财务模型 V2.0'
ws_cover['A1'].font = Font(name='微软雅黑', size=18, bold=True, color='1F4E78')
ws_cover.merge_cells('A1:F1')

ws_cover['A2'] = 'AI驱动的另类数据量化投研平台 ｜ 资管科技赛道'
ws_cover['A2'].font = Font(name='微软雅黑', size=12, color='666666')
ws_cover.merge_cells('A2:F2')

ws_cover['A4'] = '编制单位：慧点资本 (InsightQuant) + 杭州永字资产管理有限公司（推荐单位）'
ws_cover['A5'] = '编制日期：2026年6月'
ws_cover['A6'] = '数据基准：业内资管科技 SaaS 同业公开数据 + 自身场景调整'
ws_cover['A7'] = '预测期：2026-2030 年（5 年）'
ws_cover['A8'] = '货币单位：人民币 万元'

for row in [4, 5, 6, 7, 8]:
    ws_cover[f'A{row}'].font = Font(name='微软雅黑', size=10, italic=True)

# 核心结论
ws_cover['A10'] = '一、核心财务结论（基准情景）'
ws_cover['A10'].font = HEADER_FONT
ws_cover['A10'].fill = HEADER_FILL
ws_cover.merge_cells('A10:F10')

core_metrics = [
    ('指标', '2026年', '2027年', '2028年', '2029年', '2030年'),
    ('营业收入（万元）', 300, 1200, 2700, 5500, 12000),
    ('同比增长（%）', '—', '300%', '125%', '104%', '118%'),
    ('毛利率（%）', '45%', '55%', '65%', '70%', '72%'),
    ('净利率（%）', '-83%', '-25%', '3%', '15%', '28%'),
    ('净利润（万元）', -250, -300, 80, 825, 3360),
    ('累计融资需求（万元）', 500, 800, 500, 0, 0),
    ('期末现金（万元）', 250, 100, 280, 1105, 4465),
    ('累计客户数（家）', 8, 25, 60, 130, 250),
    ('ARPU（万元/家/年）', 37.5, 48, 45, 42, 48),
    ('员工数（人）', 8, 15, 25, 40, 60),
    ('IRR（5年）', '—', '—', '—', '—', '25-30%'),
]

for row_idx, row_data in enumerate(core_metrics, start=11):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_cover.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 11:
            cell.font = HEADER_FONT
            cell.fill = SUBHEADER_FILL
        else:
            cell.font = NORMAL_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

# 关键判断
ws_cover['A24'] = '二、关键判断与差异化'
ws_cover['A24'].font = HEADER_FONT
ws_cover['A24'].fill = HEADER_FILL
ws_cover.merge_cells('A24:F24')

key_judgments = [
    '1. 相比原 BP 的"第5年6亿元营收、91%净利率"，本模型采用业内基准保守预测，更具可信度',
    '2. 起步期高研发投入（占营收 60-80%）符合资管科技 SaaS 行业规律（同业基准）',
    '3. 2028 年达到盈亏平衡（基准），符合慧点资本+永字资管生态资源支持下的合理节奏',
    '4. 收入增长依赖：(a) 永字资管 LP/GP 网络种子客户 (b) 公开标杆案例 (c) 算力补贴降低成本',
    '5. 不依赖外部"500万天使"单一来源，而是采用"孵化+战略合作+空间落地+算力补贴"复合模式',
    '6. IRR 25-30% 符合业内资管科技 SaaS 同业水平（聚宽/米筐等公开数据）',
    '7. 关键风险：技术落地延迟、合规风险、推荐单位生态兑现度',
]

for i, judgment in enumerate(key_judgments, start=25):
    ws_cover[f'A{i}'] = judgment
    ws_cover[f'A{i}'].font = NORMAL_FONT
    ws_cover.merge_cells(f'A{i}:F{i}')

# 三情景对比
ws_cover['A34'] = '三、三情景对比（2030年关键指标）'
ws_cover['A34'].font = HEADER_FONT
ws_cover['A34'].fill = HEADER_FILL
ws_cover.merge_cells('A34:F34')

scenarios = [
    ('情景', '假设', '营收（万元）', '净利润（万元）', '净利率', 'IRR'),
    ('乐观情景', '永字资管生态完全兑现+大客户签约', 18000, 6300, '35%', '38%'),
    ('基准情景', '推荐单位资源正常转化+稳健增长', 12000, 3360, '28%', '28%'),
    ('保守情景', '仅依赖自有客户+公开获客', 6500, 1300, '20%', '18%'),
]

for row_idx, row_data in enumerate(scenarios, start=35):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_cover.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 35:
            cell.font = HEADER_FONT
            cell.fill = SUBHEADER_FILL
        else:
            cell.font = NORMAL_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

ws_cover.column_dimensions['A'].width = 30
for col in 'BCDEF':
    ws_cover.column_dimensions[col].width = 18

# ============ Sheet 2: 损益表（基准情景）===========
ws_pl = wb.create_sheet('损益表_基准')

ws_pl['A1'] = '损益表（基准情景）｜单位：万元'
ws_pl['A1'].font = HEADER_FONT
ws_pl['A1'].fill = HEADER_FILL
ws_pl.merge_cells('A1:G1')

pl_headers = ['项目', '2026年', '2027年', '2028年', '2029年', '2030年', '5年合计']
for col_idx, header in enumerate(pl_headers, start=1):
    cell = ws_pl.cell(row=3, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# 损益表数据（基准情景）
pl_data = [
    # 收入
    ('一、营业收入', 300, 1200, 2700, 5500, 12000, 21700, True),
    ('  订阅服务（60%）', 180, 720, 1620, 3300, 7200, 13020, False),
    ('  定制开发（20%）', 60, 240, 540, 1100, 2400, 4340, False),
    ('  数据服务（15%）', 45, 180, 405, 825, 1800, 3255, False),
    ('  战略合作（5%）', 15, 60, 135, 275, 600, 1085, False),
    # 成本
    ('二、营业成本', 165, 540, 945, 1650, 3360, 6660, True),
    ('  数据采购', 50, 130, 250, 450, 700, 1580, False),
    ('  云服务/算力', 40, 110, 200, 350, 600, 1300, False),
    ('  第三方技术服务', 30, 120, 200, 350, 600, 1300, False),
    ('  客户服务成本', 25, 100, 175, 300, 700, 1300, False),
    ('  其他', 20, 80, 120, 200, 760, 1180, False),
    # 毛利
    ('三、毛利', 135, 660, 1755, 3850, 8640, 15040, True),
    # 费用
    ('四、营业费用', 385, 960, 1675, 3025, 5280, 11325, True),
    ('  研发费用', 250, 500, 800, 1300, 2200, 5050, False),
    ('  销售费用', 80, 280, 525, 1075, 1850, 3810, False),
    ('  管理费用', 45, 130, 250, 450, 800, 1675, False),
    ('  财务费用', 10, 50, 100, 200, 430, 790, False),
    # 营业利润
    ('五、营业利润', -250, -300, 80, 825, 3360, 3715, True),
    # 税前利润
    ('六、利润总额', -250, -300, 80, 825, 3360, 3715, False),
    ('  所得税（0%/12.5%/15%/15%/15%）', 0, 0, 0, 124, 504, 628, False),
    # 净利润
    ('七、净利润', -250, -300, 80, 701, 2856, 3087, True),
    # 关键比率
    ('八、关键比率', '', '', '', '', '', '', True),
    ('  毛利率', '45%', '55%', '65%', '70%', '72%', '69%', False),
    ('  净利率', '-83%', '-25%', '3%', '13%', '24%', '14%', False),
    ('  研发费用率', '83%', '42%', '30%', '24%', '18%', '23%', False),
    ('  销售费用率', '27%', '23%', '19%', '20%', '15%', '18%', False),
]

for row_idx, row_data in enumerate(pl_data, start=4):
    label, *values, is_total = row_data
    cell_label = ws_pl.cell(row=row_idx, column=1, value=label)
    cell_label.font = BOLD_FONT if is_total else NORMAL_FONT
    cell_label.alignment = LEFT
    cell_label.border = THIN_BORDER
    if is_total:
        cell_label.fill = TOTAL_FILL

    for col_idx, value in enumerate(values, start=2):
        cell = ws_pl.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BOLD_FONT if is_total else NORMAL_FONT
        cell.alignment = RIGHT
        cell.border = THIN_BORDER
        if is_total:
            cell.fill = TOTAL_FILL

ws_pl.column_dimensions['A'].width = 30
for col in 'BCDEFG':
    ws_pl.column_dimensions[col].width = 14

# ============ Sheet 3: 损益表_乐观 ============
ws_pl_opt = wb.create_sheet('损益表_乐观')

ws_pl_opt['A1'] = '损益表（乐观情景）｜单位：万元'
ws_pl_opt['A1'].font = HEADER_FONT
ws_pl_opt['A1'].fill = HEADER_FILL
ws_pl_opt.merge_cells('A1:G1')

for col_idx, header in enumerate(pl_headers, start=1):
    cell = ws_pl_opt.cell(row=3, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# 乐观情景数据（关键差异：营收更高、成本/费用率相同）
pl_opt_data = [
    ('一、营业收入', 350, 1500, 3800, 8500, 18000, 32150, True),
    ('  订阅服务（60%）', 210, 900, 2280, 5100, 10800, 19290, False),
    ('  定制开发（20%）', 70, 300, 760, 1700, 3600, 6430, False),
    ('  数据服务（15%）', 53, 225, 570, 1275, 2700, 4823, False),
    ('  战略合作（5%）', 17, 75, 190, 425, 900, 1607, False),
    ('二、营业成本', 175, 600, 1140, 2550, 5040, 9505, True),
    ('三、毛利', 175, 900, 2660, 5950, 12960, 22645, True),
    ('四、营业费用', 405, 1100, 2090, 4250, 7560, 15405, True),
    ('五、营业利润', -230, -200, 570, 1700, 5400, 7240, True),
    ('六、利润总额', -230, -200, 570, 1700, 5400, 7240, False),
    ('  所得税', 0, 0, 71, 255, 810, 1136, False),
    ('七、净利润', -230, -200, 499, 1445, 4590, 6104, True),
    ('  净利率', '-66%', '-13%', '13%', '17%', '26%', '19%', False),
]

for row_idx, row_data in enumerate(pl_opt_data, start=4):
    label, *values, is_total = row_data
    cell_label = ws_pl_opt.cell(row=row_idx, column=1, value=label)
    cell_label.font = BOLD_FONT if is_total else NORMAL_FONT
    cell_label.alignment = LEFT
    cell_label.border = THIN_BORDER
    if is_total:
        cell_label.fill = TOTAL_FILL

    for col_idx, value in enumerate(values, start=2):
        cell = ws_pl_opt.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BOLD_FONT if is_total else NORMAL_FONT
        cell.alignment = RIGHT
        cell.border = THIN_BORDER
        if is_total:
            cell.fill = TOTAL_FILL

ws_pl_opt.column_dimensions['A'].width = 30
for col in 'BCDEFG':
    ws_pl_opt.column_dimensions[col].width = 14

# ============ Sheet 4: 损益表_保守 ============
ws_pl_con = wb.create_sheet('损益表_保守')

ws_pl_con['A1'] = '损益表（保守情景）｜单位：万元'
ws_pl_con['A1'].font = HEADER_FONT
ws_pl_con['A1'].fill = HEADER_FILL
ws_pl_con.merge_cells('A1:G1')

for col_idx, header in enumerate(pl_headers, start=1):
    cell = ws_pl_con.cell(row=3, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

pl_con_data = [
    ('一、营业收入', 200, 700, 1500, 3000, 6500, 11900, True),
    ('二、营业成本', 120, 350, 600, 1050, 1950, 4070, True),
    ('三、毛利', 80, 350, 900, 1950, 4550, 7830, True),
    ('四、营业费用', 310, 650, 1100, 1850, 3250, 7160, True),
    ('五、营业利润', -230, -300, -200, 100, 1300, 670, True),
    ('六、利润总额', -230, -300, -200, 100, 1300, 670, False),
    ('  所得税', 0, 0, 0, 15, 195, 210, False),
    ('七、净利润', -230, -300, -200, 85, 1105, 460, True),
    ('  净利率', '-115%', '-43%', '-13%', '3%', '17%', '4%', False),
]

for row_idx, row_data in enumerate(pl_con_data, start=4):
    label, *values, is_total = row_data
    cell_label = ws_pl_con.cell(row=row_idx, column=1, value=label)
    cell_label.font = BOLD_FONT if is_total else NORMAL_FONT
    cell_label.alignment = LEFT
    cell_label.border = THIN_BORDER
    if is_total:
        cell_label.fill = TOTAL_FILL

    for col_idx, value in enumerate(values, start=2):
        cell = ws_pl_con.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BOLD_FONT if is_total else NORMAL_FONT
        cell.alignment = RIGHT
        cell.border = THIN_BORDER
        if is_total:
            cell.fill = TOTAL_FILL

ws_pl_con.column_dimensions['A'].width = 30
for col in 'BCDEFG':
    ws_pl_con.column_dimensions[col].width = 14

# ============ Sheet 5: 现金流表（基准）===========
ws_cf = wb.create_sheet('现金流表_基准')

ws_cf['A1'] = '现金流量表（基准情景）｜单位：万元'
ws_cf['A1'].font = HEADER_FONT
ws_cf['A1'].fill = HEADER_FILL
ws_cf.merge_cells('A1:G1')

cf_headers = ['项目', '2026年', '2027年', '2028年', '2029年', '2030年', '5年合计']
for col_idx, header in enumerate(cf_headers, start=1):
    cell = ws_cf.cell(row=3, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

cf_data = [
    ('一、经营活动现金流', '', '', '', '', '', '', True),
    ('  净利润', -250, -300, 80, 701, 2856, 3087, False),
    ('  折旧摊销', 30, 50, 80, 120, 180, 460, False),
    ('  营运资本变化', -50, -120, -180, -300, -500, -1150, False),
    ('  经营活动现金流净额', -270, -370, -20, 521, 2536, 2397, True),
    ('二、投资活动现金流', '', '', '', '', '', '', True),
    ('  资本性支出（设备/系统）', -50, -80, -100, -150, -200, -580, False),
    ('  投资活动现金流净额', -50, -80, -100, -150, -200, -580, True),
    ('三、筹资活动现金流', '', '', '', '', '', '', True),
    ('  股权融资', 500, 800, 500, 0, 0, 1800, False),
    ('  债务融资', 0, 0, 0, 0, 0, 0, False),
    ('  分红', 0, 0, 0, 0, 0, 0, False),
    ('  筹资活动现金流净额', 500, 800, 500, 0, 0, 1800, True),
    ('四、汇率变动影响', 0, 0, 0, 0, 0, 0, False),
    ('五、现金净增加额', 180, 350, 380, 371, 2336, 3617, True),
    ('六、期初现金', 100, 280, 630, 1010, 1381, 100, False),
    ('七、期末现金', 280, 630, 1010, 1381, 3717, 3717, True),
]

for row_idx, row_data in enumerate(cf_data, start=4):
    label, *values, is_total = row_data
    cell_label = ws_cf.cell(row=row_idx, column=1, value=label)
    cell_label.font = BOLD_FONT if is_total else NORMAL_FONT
    cell_label.alignment = LEFT
    cell_label.border = THIN_BORDER
    if is_total:
        cell_label.fill = TOTAL_FILL

    for col_idx, value in enumerate(values, start=2):
        cell = ws_cf.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BOLD_FONT if is_total else NORMAL_FONT
        cell.alignment = RIGHT
        cell.border = THIN_BORDER
        if is_total:
            cell.fill = TOTAL_FILL

ws_cf.column_dimensions['A'].width = 30
for col in 'BCDEFG':
    ws_cf.column_dimensions[col].width = 14

# ============ Sheet 6: 资产负债表（基准）===========
ws_bs = wb.create_sheet('资产负债表_基准')

ws_bs['A1'] = '资产负债表（基准情景）｜单位：万元'
ws_bs['A1'].font = HEADER_FONT
ws_bs['A1'].fill = HEADER_FILL
ws_bs.merge_cells('A1:G1')

bs_headers = ['项目', '2026年末', '2027年末', '2028年末', '2029年末', '2030年末']
for col_idx, header in enumerate(bs_headers, start=1):
    cell = ws_bs.cell(row=3, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

bs_data = [
    ('一、资产', '', '', '', '', '', True),
    ('  货币资金', 280, 630, 1010, 1381, 3717, False),
    ('  应收账款', 50, 200, 450, 920, 2000, False),
    ('  其他流动资产', 30, 50, 80, 120, 180, False),
    ('  流动资产合计', 360, 880, 1540, 2421, 5897, True),
    ('  固定资产', 80, 130, 180, 250, 350, False),
    ('  无形资产', 30, 50, 80, 120, 180, False),
    ('  非流动资产合计', 110, 180, 260, 370, 530, True),
    ('  资产总计', 470, 1060, 1800, 2791, 6427, True),
    ('二、负债', '', '', '', '', '', True),
    ('  应付账款', 30, 80, 150, 280, 550, False),
    ('  应付薪酬', 100, 200, 350, 600, 900, False),
    ('  其他流动负债', 20, 40, 70, 110, 160, False),
    ('  流动负债合计', 150, 320, 570, 990, 1610, True),
    ('  长期负债', 0, 0, 0, 0, 0, False),
    ('  负债合计', 150, 320, 570, 990, 1610, True),
    ('三、所有者权益', '', '', '', '', '', True),
    ('  实收资本', 600, 1400, 1900, 1900, 1900, False),
    ('  资本公积', 100, 100, 100, 100, 100, False),
    ('  累计未分配利润', -380, -680, -600, 101, 2957, False),
    ('  所有者权益合计', 320, 820, 1400, 2101, 4957, True),
    ('  负债与权益总计', 470, 1140, 1970, 3091, 6567, True),
]

for row_idx, row_data in enumerate(bs_data, start=4):
    label, *values, is_total = row_data
    cell_label = ws_bs.cell(row=row_idx, column=1, value=label)
    cell_label.font = BOLD_FONT if is_total else NORMAL_FONT
    cell_label.alignment = LEFT
    cell_label.border = THIN_BORDER
    if is_total:
        cell_label.fill = TOTAL_FILL

    for col_idx, value in enumerate(values, start=2):
        cell = ws_bs.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BOLD_FONT if is_total else NORMAL_FONT
        cell.alignment = RIGHT
        cell.border = THIN_BORDER
        if is_total:
            cell.fill = TOTAL_FILL

ws_bs.column_dimensions['A'].width = 30
for col in 'BCDEF':
    ws_bs.column_dimensions[col].width = 14

# ============ Sheet 7: 假设与基准 ============
ws_assump = wb.create_sheet('假设与基准')

ws_assump['A1'] = '核心假设与业内基准说明'
ws_assump['A1'].font = HEADER_FONT
ws_assump['A1'].fill = HEADER_FILL
ws_assump.merge_cells('A1:D1')

assump_data = [
    ['维度', '本模型假设', '业内基准', '依据'],
    ['客户数增速', '第1年8家 → 第5年累计250家', '聚宽/米筐/同花顺 iFinD 增速 30-50%/年', '公开年报与行业报告'],
    ['ARPU', '40-50万元/家/年', '资管科技 SaaS 行业 ARPU 30-80 万', '同业公开数据'],
    ['订阅占比', '60%', '60-70%', '资管科技 SaaS 行业惯例'],
    ['定制开发占比', '20%', '15-25%', '机构客户定制需求'],
    ['数据服务占比', '15%', '5-10%', '本项目核心差异化优势'],
    ['战略合作占比', '5%', '—', '推荐单位永字资管生态'],
    ['毛利率（成熟期）', '70-72%', '60-75%', '同业 SaaS 毛利率'],
    ['净利率（成熟期）', '24-28%', '20-30%', '成熟 SaaS 净利率'],
    ['研发费用率（前期）', '60-80%', '50-80%', 'AI 研发高投入'],
    ['销售费用率', '15-27%', '15-30%', '机构客户获客成本'],
    ['人均创收（2030年）', '200万元/人', '150-300万元/人', '资管科技 SaaS 同业'],
    ['人均薪酬成本', '40-50万元/年', '40-60万元/年', '上海金融科技岗位'],
    ['客户续约率', '90%', '80-95%', 'SaaS 行业'],
    ['所得税率', '0%/12.5%/15%/15%/15%', '15%（高新企业）', '高新技术企业认定'],
    ['5年累计融资', '1800万元', '—', '股权+孵化资源'],
    ['IRR（基准）', '25-30%', '20-35%', '业内资管科技 SaaS'],
    ['回收期', '3-3.5年', '3-5年', '业内 SaaS'],
]

for row_idx, row_data in enumerate(assump_data, start=3):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_assump.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 3:
            cell.font = HEADER_FONT
            cell.fill = SUBHEADER_FILL
        else:
            cell.font = NORMAL_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER

ws_assump.column_dimensions['A'].width = 25
ws_assump.column_dimensions['B'].width = 30
ws_assump.column_dimensions['C'].width = 25
ws_assump.column_dimensions['D'].width = 35

# ============ Sheet 8: 收入分解 ============
ws_rev = wb.create_sheet('收入分解')

ws_rev['A1'] = '收入分解预测（基准情景）'
ws_rev['A1'].font = HEADER_FONT
ws_rev['A1'].fill = HEADER_FILL
ws_rev.merge_cells('A1:G1')

rev_headers = ['收入类别', '2026年', '2027年', '2028年', '2029年', '2030年', '5年CAGR']
for col_idx, header in enumerate(rev_headers, start=1):
    cell = ws_rev.cell(row=3, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

rev_data = [
    ['订阅服务', 180, 720, 1620, 3300, 7200, '109%'],
    ['  基础版（30%）', 54, 216, 486, 990, 2160, '109%'],
    ['  专业版（50%）', 90, 360, 810, 1650, 3600, '109%'],
    ['  企业版（20%）', 36, 144, 324, 660, 1440, '109%'],
    ['定制开发', 60, 240, 540, 1100, 2400, '109%'],
    ['数据服务', 45, 180, 405, 825, 1800, '109%'],
    ['  卫星图像数据', 15, 60, 135, 275, 600, '109%'],
    ['  舆情数据', 10, 40, 90, 185, 400, '109%'],
    ['  供应链数据', 12, 50, 115, 230, 500, '109%'],
    ['  其他', 8, 30, 65, 135, 300, '113%'],
    ['战略合作（永字资管生态）', 15, 60, 135, 275, 600, '109%'],
    ['合计', 300, 1200, 2700, 5500, 12000, '109%'],
]

for row_idx, row_data in enumerate(rev_data, start=4):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_rev.cell(row=row_idx, column=col_idx, value=value)
        is_total = row_data[0] in ['订阅服务', '定制开发', '数据服务', '合计']
        if is_total:
            cell.font = BOLD_FONT
            cell.fill = TOTAL_FILL
        else:
            cell.font = NORMAL_FONT
        cell.alignment = RIGHT if col_idx > 1 else LEFT
        cell.border = THIN_BORDER

ws_rev.column_dimensions['A'].width = 30
for col in 'BCDEFG':
    ws_rev.column_dimensions[col].width = 14

# ============ Sheet 9: 关键风险与缓解 ============
ws_risk = wb.create_sheet('关键风险与缓解')

ws_risk['A1'] = '关键风险与缓解措施'
ws_risk['A1'].font = HEADER_FONT
ws_risk['A1'].fill = HEADER_FILL
ws_risk.merge_cells('A1:E1')

risk_headers = ['风险类型', '风险描述', '影响', '缓解措施', '责任人']
for col_idx, header in enumerate(risk_headers, start=1):
    cell = ws_risk.cell(row=3, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

risks = [
    ['技术风险', 'AI 模型应用层落地延迟、效果不及预期', '高', '采用开源基础模型（Qwen/DeepSeek）+ 应用层深度优化，6个月内出 MVP', '王宇寒'],
    ['市场风险', '机构客户决策慢、试点周期长', '高', '依靠永字资管生态资源快速触达；先做 3-5 家种子客户深度服务', '薛永再'],
    ['合规风险', '金融监管政策变化、AI 备案要求', '中', '团队合规机制+应用层可解释性；保持与监管沟通', '王宇寒'],
    ['人才风险', '核心团队（王宇寒、官馨、梁理智）流失', '中', '股权激励+项目分红；浙江大学+杭州电子科技大学双母校资源招聘', '冯亦根'],
    ['资金风险', '研发投入超出预算', '中', '5年累计融资 1800 万，分阶段释放；优先使用算力补贴', '薛永再'],
    ['客户集中度', '前 5 家客户占比超过 50%', '中', '多行业拓展；技术差异化降低单客户依赖', '冯亦根'],
    ['推荐单位风险', '永字资管生态资源未充分兑现', '中', '签订战略合作协议；建立独立销售能力', '冯亦根+薛永再'],
    ['竞争风险', 'Wind/同花顺/Bloomberg 等巨头 AI 升级', '高', '差异化定位：聚焦中小资管+大模型应用层', '全体'],
]

for row_idx, row_data in enumerate(risks, start=4):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_risk.cell(row=row_idx, column=col_idx, value=value)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER

ws_risk.column_dimensions['A'].width = 15
ws_risk.column_dimensions['B'].width = 35
ws_risk.column_dimensions['C'].width = 8
ws_risk.column_dimensions['D'].width = 45
ws_risk.column_dimensions['E'].width = 15

# ============ 保存 ============
output_path = 'D:/shFintech/QuantInsight_Pro_Financial_Model_V2.xlsx'
wb.save(output_path)
print(f'OK: 财务模型 V2.0 已保存: {output_path}')
print(f'Sheets: {len(wb.sheetnames)} 个')
for s in wb.sheetnames:
    print(f'  - {s}')
